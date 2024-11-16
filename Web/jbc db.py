import os
import re
import time
import logging
import glob
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import win32com.client as win32
import mysql.connector  # 用于数据库连接

# 数据库配置
db_config = {
    'host': 'localhost',
    'user': 'root',
    'password': '',
    'database': 'osaka_main'
}

# 设置 JBC 文件夹路径
jbc_folder = r"C:\Server\JBC"

# 确保目录存在
if not os.path.exists(jbc_folder):
    os.makedirs(jbc_folder)


def get_db_connection():
    """建立数据库连接"""
    return mysql.connector.connect(**db_config)


def query_database_for_single_numbers():
    """从数据库中查询需要的单号列表"""
    connection = get_db_connection()
    cursor = connection.cursor()
    
    query = """
    SELECT DISTINCT single_number 
    FROM shipments 
    WHERE status = 'NAKAMURA' OR status = 'GB';
    """
    cursor.execute(query)
    single_numbers = [row[0] for row in cursor.fetchall()]
    
    cursor.close()
    connection.close()
    
    print(f"查询到的单号：{single_numbers}")
    return single_numbers


def download_csv(username, password):
    download_path = os.path.expanduser('~\Server\JBC')
    chrome_options = Options()
    prefs = {
        'download.default_directory': download_path,
        'download.prompt_for_download': False,
        'profile.default_content_settings.popups': 0,
        'directory_upgrade': True
    }
    chrome_options.add_experimental_option('prefs', prefs)
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--window-size=1920x1080')

    driver = webdriver.Chrome(options=chrome_options)

    try:
        print(f"开始下载用户 {username} 的 CSV 文件")
        driver.get("http://www.ankcustoms.com/login.aspx")
        username_input = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "TextBox1"))
        )
        username_input.send_keys(username)
        password_input = driver.find_element(By.ID, "TextBox2")
        password_input.send_keys(password)
        driver.find_element(By.ID, "Button1").click()
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "rdoTrans"))
        ).click()

        start_datetime = (datetime.now() - timedelta(days=2)).strftime("%Y-%m-%d %H:%M:%S")
        end_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        driver.find_element(By.ID, "_easyui_textbox_input5").send_keys(start_datetime)
        driver.find_element(By.ID, "_easyui_textbox_input6").send_keys(end_datetime)
        time.sleep(2)
        driver.find_element(By.CSS_SELECTOR, "a.easyui-linkbutton[data-options*='icon-save']").click()
        print("等待下载完成...")
        wait_for_download_completion(download_path)

    finally:
        driver.quit()


def wait_for_download_completion(download_path, timeout=60):
    initial_files = set(glob.glob(f"{download_path}/*.csv"))
    for _ in range(timeout):
        time.sleep(1)
        current_files = set(glob.glob(f"{download_path}/*.csv"))
        new_files = current_files - initial_files
        if new_files:
            return list(new_files)[0]
    print("下载超时")
    return None


def process_csv_files_using_database(csv_files):
    """处理 CSV 文件，并使用数据库中的数据匹配"""
    dfs = [pd.read_csv(f, encoding='utf-8') for f in csv_files]
    merged_df = pd.concat(dfs, ignore_index=True).dropna()
    merged_df.iloc[:, 2] = merged_df.iloc[:, 2].astype(str).str.replace("-", "")

    single_numbers = query_database_for_single_numbers()
    filtered_df = merged_df[merged_df.iloc[:, 2].isin(single_numbers)]

    filtered_filename = datetime.now().strftime("%Y-%m-%d_filtered.csv")
    filtered_df.to_csv(filtered_filename, index=False)
    print(f"Filtered CSV file saved as {filtered_filename}")
    return filtered_filename


def get_single_numbers(filtered_filename):
    filtered_data = pd.read_csv(filtered_filename)
    return filtered_data.iloc[:, 2].tolist()


def download_uof_file():
    download_path = r"C:\Server\JBC"
    # 此部分代码保持不变
    # 省略其他代码...（因篇幅限制，实际代码保持不变）
    pass


def clear_old_uof_files(download_path):
    old_files = glob.glob(os.path.join(download_path, "UOF出入库汇总表*.xlsx"))
    for file in old_files:
        try:
            os.remove(file)
            print(f"已删除旧文件：{file}")
        except Exception as e:
            print(f"删除文件 {file} 时出错：{e}")


def create_excel_file(single_numbers, uof_data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # 设置列名称和宽度
    column_names = ["许可时间", "回数", "送り状番号", "箱数", "转运公司", "转运备注", "现场用-函数对应", "入库时间", "取件地"]
    column_widths = [11, 4.63, 13.25, 4.63, 18, 25, 58.13, 19.88, 15]
    sheet.append(column_names)
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width

    current_date = datetime.now().strftime("%Y/%m/%d")

    # 填充数据
    for single_number in single_numbers:
        print(f"正在写入单号: {single_number}")
        match = uof_data[uof_data['送り状番号'].astype(str).str.contains(str(single_number), na=False)]
        if not match.empty:
            matched_row = match.iloc[0]
            sheet.append([current_date, "", single_number, matched_row['箱数'], matched_row['转运公司'], matched_row['转运备注'], matched_row['现场用-函数对应'], matched_row['入库时间'], matched_row['取件地']])
        else:
            sheet.append([current_date, "", single_number, "", "", "", "", "", ""])
        print(f"写入完成单号: {single_number}")

    # 保存 Excel 文件
    date_string = datetime.now().strftime('%Y-%m-%d-%H%M')
    excel_filename = f"{date_string}.xlsx"
    file_path = os.path.join(jbc_folder, excel_filename)
    workbook.save(file_path)
    print(f"Excel 文件已创建并保存: {file_path}")
    return file_path


def format_excel_file(excel_filename):
    try:
        if not os.path.isfile(excel_filename):
            print(f"Error: File not found: {excel_filename}")
            return

        Excel = win32.DispatchEx("Excel.Application")
        Excel.Visible = False

        Workbook = Excel.Workbooks.Open(excel_filename)
        Worksheet = Workbook.Worksheets(1)

        for row in range(1, Worksheet.UsedRange.Rows.Count + 1):
            Worksheet.Rows(row).RowHeight = 30

        for col in range(1, 10):
            if col != 7:
                Worksheet.Columns(col).HorizontalAlignment = -4108
                Worksheet.Columns(col).VerticalAlignment = -4108
        Worksheet.Columns("G").VerticalAlignment = -4108

        margin_points = 0.9 * 28.35
        Worksheet.PageSetup.TopMargin = margin_points
        Worksheet.PageSetup.BottomMargin = margin_points

        for i in range(Worksheet.UsedRange.Rows.Count, 1, -1):
            if Worksheet.Cells(i, 8).Value is not None:
                Worksheet.Rows(i).Delete()

        thin_border = 2
        for row in range(1, Worksheet.UsedRange.Rows.Count + 1):
            if Worksheet.Cells(row, 1).Value is not None:
                for col in range(1, 10):
                    Worksheet.Cells(row, col).Borders.Weight = thin_border

        for row in range(1, Worksheet.UsedRange.Rows.Count + 1):
            if row % 2 == 0:
                Worksheet.Rows(row).Interior.Color = 0xFFFFFF
            else:
                Worksheet.Rows(row).Interior.Color = 0xF2F2F2

        Worksheet.PageSetup.CenterFooter = "第&P页, 共&N页"
        Worksheet.PageSetup.PrintArea = f'A1:I{Worksheet.UsedRange.Rows.Count}'
        Worksheet.PageSetup.Orientation = 2
        Worksheet.PageSetup.PaperSize = 9
        Worksheet.PageSetup.Zoom = False
        Worksheet.PageSetup.FitToPagesWide = 1
        Worksheet.PageSetup.FitToPagesTall = False

        Workbook.Save()
        print("Excel 文件已保存。")

        date_string = datetime.now().strftime('%Y-%m-%d-%H%M')
        output_filename = f"{date_string}-许可.pdf"
        output_path = os.path.join(jbc_folder, output_filename)

        time.sleep(2)

        try:
            Workbook.ExportAsFixedFormat(0, output_path)
            print(f"PDF 文件已保存: {output_path}")
        except Exception as e:
            print(f"Error exporting to PDF: {e}")

        Workbook.Close(SaveChanges=False)

    except Exception as e:
        logging.exception("An error occurred:")
        print(f"An error occurred: {str(e).encode('utf-8', errors='ignore').decode('utf-8')}")

    finally:
        if Excel:
            Excel.Quit()
        print("Excel 进程已退出。")


def main():
    csv_files = [
        download_csv("JUTB", "JUTp&HJKL2SJYjjuutt"),
        download_csv("UOFB", "EWQ&6qwe42B")
    ]
    csv_files = [f for f in csv_files if f]
    if not csv_files:
        print("CSV 文件下载失败。程序退出。")
        return

    filtered_filename = process_csv_files_using_database(csv_files)
    single_numbers = get_single_numbers(filtered_filename)

    uof_file = download_uof_file()
    if uof_file:
        try:
            print(f"读取 UOF 文件：{uof_file}")
            uof_data = pd.read_excel(uof_file, engine='openpyxl')
        except Exception as e:
            print(f"读取 UOF 文件时出错：{e}")
            return
    else:
        print("UOF 文件下载失败。程序退出。")
        return

    excel_filename = create_excel_file(single_numbers, uof_data)
    format_excel_file(excel_filename)


if __name__ == "__main__":
    main()
