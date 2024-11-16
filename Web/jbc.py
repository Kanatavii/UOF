import os
import re
import time
import logging
import glob
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import win32com.client as win32

# 设置 JBC 文件夹路径
jbc_folder = r"C:\Server\JBC"

# 确保目录存在
if not os.path.exists(jbc_folder):
    os.makedirs(jbc_folder)
    
def download_csv(username, password):
    download_path = os.path.expanduser('~\\Server\JBC')
    chrome_options = Options()
    prefs = {
        'download.default_directory': download_path,
        'download.prompt_for_download': False,
        'profile.default_content_settings.popups': 0,
        'directory_upgrade': True
    }
    chrome_options.add_experimental_option('prefs', prefs)
    # 可选：使用无头模式以加快执行速度
    # chrome_options.add_argument('--headless')
    # 添加无头模式选项
    chrome_options.add_argument('--headless')  # 运行在后台的无头模式
    chrome_options.add_argument('--disable-gpu')  # 禁用GPU加速
    chrome_options.add_argument('--no-sandbox')  # 仅在Linux上运行时需要
    chrome_options.add_argument('--window-size=1920x1080')  # 设置窗口大小
    
    driver = webdriver.Chrome(options=chrome_options)
    
    try:
        print(f"开始下载用户 {username} 的 CSV 文件")

        # 访问登录页面
        driver.get("http://www.ankcustoms.com/login.aspx")
        print("已打开登录页面")

        # 登录
        username_input = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "TextBox1"))
        )
        username_input.send_keys(username)
        print("已输入用户名")

        password_input = driver.find_element(By.ID, "TextBox2")
        password_input.send_keys(password)
        print("已输入密码")

        login_button = driver.find_element(By.ID, "Button1")
        login_button.click()
        print("已点击登录按钮")

        # 等待页面加载
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "rdoTrans"))
        )
        print("登录成功，页面已加载")

        # 选择“根据转运时间查询”单选按钮
        radio_button = driver.find_element(By.ID, "rdoTrans")
        radio_button.click()
        print("已选择根据转运时间查询")

        # 设置起始和结束日期时间
        start_datetime = (datetime.now() - timedelta(days=2)).strftime("%Y-%m-%d %H:%M:%S")
        end_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"开始时间：{start_datetime}，结束时间：{end_datetime}")

        # 定位输入框
        start_input = driver.find_element(By.ID, "_easyui_textbox_input5")
        end_input = driver.find_element(By.ID, "_easyui_textbox_input6")

        # 输入日期和时间
        start_input.clear()
        start_input.send_keys(start_datetime)
        print("已输入开始时间")

        end_input.clear()
        end_input.send_keys(end_datetime)
        print("已输入结束时间")

        # 等待页面处理输入
        time.sleep(2)

        # 点击“CSV下载”链接
        csv_download_link = driver.find_element(By.CSS_SELECTOR, "a.easyui-linkbutton[data-options*='icon-save']")
        csv_download_link.click()
        print("已点击 CSV 下载链接")

        # 等待下载完成
        print("等待下载完成...")
        wait_for_download_completion(download_path)
        print("下载完成")

        # 获取下载的 CSV 文件
        downloaded_file = get_latest_downloaded_file(download_path)
        print(f"下载的 CSV 文件：{downloaded_file}")

    except Exception as e:
        logging.exception(f"Error downloading CSV for user {username}: {e}")
        print(f"Error downloading CSV for user {username}: {e}")
        downloaded_file = None
    finally:
        print("关闭浏览器...")
        driver.quit()
        print("浏览器已关闭")

    return downloaded_file

def wait_for_download_completion(download_path=r"C:\Server\JBC", timeout=60):
    """等待 CSV 文件下载完成"""
    if isinstance(timeout, str):
        timeout = int(timeout)  # 确保 timeout 是整数

    # 获取下载前的文件列表（使用绝对路径）
    initial_csv_files = set(glob.glob(f"{download_path}\\*.csv"))
    seconds = 0

    while seconds < timeout:
        time.sleep(1)

        # 获取当前目录中的 CSV 文件列表（使用绝对路径）
        current_files = set(glob.glob(f"{download_path}\\*.csv"))
        new_files = current_files - initial_csv_files  # 计算新下载的文件

        if new_files:
            latest_file = list(new_files)[0]  # 获取新文件中的第一个文件
            print(f"下载完成的 CSV 文件：{latest_file}")
            return latest_file  # 返回找到的 CSV 文件的绝对路径

        print(f"等待 CSV 文件下载中... {seconds}秒")
        seconds += 1

    print("下载未在超时时间内完成")
    return None  # 超时后返回 None

def get_latest_downloaded_file(download_path):
    downloaded_files = glob.glob(os.path.join(download_path, "*.csv"))
    if not downloaded_files:
        raise ValueError("在下载文件夹中未找到 CSV 文件。")
    latest_file = max(downloaded_files, key=os.path.getctime)
    return latest_file


def process_csv_files(csv_files):
    dfs = [pd.read_csv(f, encoding='utf-8') for f in csv_files]
    merged_df = pd.concat(dfs, ignore_index=True).dropna()
    merged_df.iloc[:, 2] = merged_df.iloc[:, 2].astype(str).str.replace("-", "")
    filtered_df = merged_df[merged_df.iloc[:, 11].str.contains("NAKAMURA|GB", case=False, na=False)]

    filtered_filename = datetime.now().strftime("%Y-%m-%d_filtered.csv")
    filtered_df.to_csv(filtered_filename, index=False)
    print(f"Filtered CSV file saved as {filtered_filename}")
    return filtered_filename

def get_single_numbers(filtered_filename):
    filtered_data = pd.read_csv(filtered_filename)
    return filtered_data.iloc[:, 2].tolist()

def download_uof_file():
    download_path = r"C:\Server\JBC"

    # Create download path if it doesn't exist
    if not os.path.exists(download_path):
        os.makedirs(download_path)

    chrome_options = Options()
    prefs = {
        'download.default_directory': download_path,
        'download.prompt_for_download': False,
        'profile.default_content_settings.popups': 0,
        'directory_upgrade': True,
        'safebrowsing.enabled': True
    }
    chrome_options.add_experimental_option('prefs', prefs)

    # Remove headless for debugging purposes
    # chrome_options.add_argument('--headless')  # Run in headless mode
    chrome_options.add_argument('--disable-gpu')  # Disable GPU acceleration
    chrome_options.add_argument('--no-sandbox')  # Needed for Linux only
    chrome_options.add_argument('--window-size=1920x1080')  # Set window size

    driver = webdriver.Chrome(options=chrome_options)
    latest_uof_file = None  # Initialize variable

    try:
        # Initialize Chrome browser
        driver.get("https://quickconnect.to/")
        print("Opened QuickConnect website.")

        # Find the QuickConnect ID input box and enter information
        input_box = driver.find_element(By.ID, "input-id")
        input_box.send_keys("uof-jp")
        print("Entered QuickConnect ID.")

        # Use explicit wait to ensure the submit button is clickable
        wait = WebDriverWait(driver, 60)
        submit_button = wait.until(EC.element_to_be_clickable((By.ID, "input-submit")))
        submit_button.click()
        print("Submitted QuickConnect ID.")

        # Find the username input box and enter information
        username_box = wait.until(EC.presence_of_element_located((By.NAME, "username")))
        username_box.send_keys("anguri")
        login_button = driver.find_element(By.XPATH, "//div[contains(@class, 'login-btn-spinner-wrapper')]")
        login_button.click()
        print("Entered username and clicked login.")

        # Find the password input box and enter information
        password_box = wait.until(EC.presence_of_element_located((By.NAME, "current-password")))
        password_box.send_keys("uofjpA-1")
        login_button = driver.find_element(By.XPATH, "//div[contains(@class, 'login-btn-spinner-wrapper')]")
        login_button.click()
        print("Entered password and clicked login.")

        # Click on File Station
        first_element = wait.until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="sds-desktop-shortcut"]/div/li[1]/div[1]'))
        )
        driver.execute_script("arguments[0].click();", first_element)
        print("Clicked on File Station.")

        # Double-click on the UOF folder
        action = ActionChains(driver)
        uof_folder = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[contains(text(), 'UOF')]")))
        action.double_click(uof_folder).perform()
        print("Double-clicked on UOF folder.")

        # Double-click on the Transfer Data folder
        div_element = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[contains(text(),'转运数据')]")))
        action.double_click(div_element).perform()
        print("Double-clicked on Transfer Data folder.")

        # Wait for file size elements to appear
        file_size_elements = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.x-grid3-cell-inner.x-grid3-col-filesize")))
        print(f"Found {len(file_size_elements)} file size elements.")

        # If no file size elements found, print message and exit
        if not file_size_elements:
            print("No file sizes found")
            return None

        # Desired filename to download
        desired_filename = "UOF出入库汇总表.xlsx"

        # Locate all elements containing the filename information
        file_elements = driver.find_elements(By.XPATH, f"//div[contains(text(), '{desired_filename}')]")

        # Filter out elements that contain "$" (temporary files)
        filtered_file_elements = [element for element in file_elements if "$" not in element.text]

        # Select the first matching file
        desired_file_element = filtered_file_elements[0] if filtered_file_elements else None

        # Download the specified file
        if desired_file_element:
            print(f"Attempting to download the desired file: {desired_filename}...")
            action.double_click(desired_file_element).perform()

            # Wait for download to complete
            latest_uof_file = wait_for_uof_file_completion(download_path)
            if latest_uof_file:
                print("Download Successful!")
            else:
                print("Download did not complete within the expected time.")
        else:
            print(f"File {desired_filename} not found.")

            # Pause execution for debugging purposes
            time.sleep(10)

    except TimeoutException as e:
        logging.exception("Timeout while waiting for an element.")
        print("Timeout occurred: ", e)

    except Exception as e:
        logging.exception(f"Error downloading UOF file: {e}")
        print(f"Error: {e}")

    finally:
        print("Waiting to ensure download completion...")
        time.sleep(25)  # 确保下载完成的延时
        driver.quit()
        print("Driver quit.")

    return latest_uof_file

def wait_for_uof_file_completion(download_path, timeout=12000):
    """等待 UOF 文件 (.xlsx) 下载完成"""
    seconds = 0
    pattern = re.compile(r'UOF出入库汇总表(?: \((\d+)\))?\.xlsx')  # 匹配 UOF 出入库汇总表 文件名

    while seconds < timeout:
        time.sleep(1)
        current_files = os.listdir(download_path)

        # 查找符合条件的 .xlsx 文件，且不包括 .crdownload 文件
        uof_files = [f for f in current_files if f.endswith('.xlsx') and pattern.match(f)]

        if uof_files:
            # 根据创建时间找到最新的文件
            latest_uof_file = max(uof_files, key=lambda f: os.path.getctime(os.path.join(download_path, f)))
            print(f"找到最新的 UOF 文件：{latest_uof_file}")
            return os.path.join(download_path, latest_uof_file)

        print(f"等待 UOF 文件下载中... {seconds}秒")
        seconds += 1

    print("下载未在超时时间内完成")
    return None

def clear_old_uof_files(download_path):
    """清理下载目录中旧的 UOF 文件"""
    old_files = glob.glob(os.path.join(download_path, "UOF出入库汇总表*.xlsx"))
    for file in old_files:
        try:
            os.remove(file)
            print(f"已删除旧文件：{file}")
        except Exception as e:
            print(f"删除文件 {file} 时出错：{e}")

def create_excel_file(single_numbers, uof_data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set column names and widths
    column_names = ["许可时间", "回数", "送り状番号", "箱数", "转运公司", "转运备注", "现场用-函数对应", "入库时间", "取件地"]
    column_widths = [11, 4.63, 13.25, 4.63, 18, 25, 58.13, 19.88, 15]
    sheet.append(column_names)
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width

    current_date = datetime.now().strftime("%Y/%m/%d")

    # Fill data
    for single_number in single_numbers:
        print(f"正在写入单号: {single_number}")
        match = uof_data[uof_data['送り状番号'].astype(str).str.contains(str(single_number), na=False)]
        if not match.empty:
            matched_row = match.iloc[0]
            sheet.append([current_date, "", single_number, matched_row['箱数'], matched_row['转运公司'], matched_row['转运备注'], matched_row['现场用-函数对应'], matched_row['入库时间'], matched_row['取件地']])
        else:
            sheet.append([current_date, "", single_number, "", "", "", "", "", ""])
        print(f"写入完成单号: {single_number}")

    # 保存 Excel 文件，修改命名规则
    date_string = datetime.now().strftime('%Y-%m-%d-%H%M')
    excel_filename = f"{date_string}.xlsx"
    file_path = os.path.join(jbc_folder, excel_filename)  # 将文件保存在指定的 JBC 目录
    workbook.save(file_path)
    print(f"Excel 文件已创建并保存: {file_path}")
    return file_path  # 返回完整的文件路径

import os
import time
import win32com.client as win32
import logging
from datetime import datetime

# 定义 JBC 文件夹的绝对路径
jbc_folder = r"C:\Server\JBC"

def format_excel_file(excel_filename):
    try:
        # 检查 Excel 文件是否存在
        if not os.path.isfile(excel_filename):
            print(f"Error: File not found: {excel_filename}")
            return

        # 初始化 Excel 应用程序
        Excel = win32.DispatchEx("Excel.Application")
        Excel.Visible = False  # 后台运行

        # 打开 Excel 文件
        Workbook = Excel.Workbooks.Open(excel_filename)
        Worksheet = Workbook.Worksheets(1)

        # 设置行高为 30
        for row in range(1, Worksheet.UsedRange.Rows.Count + 1):
            Worksheet.Rows(row).RowHeight = 30

        # 设置对齐方式
        for col in range(1, 10):
            if col != 7:  # 跳过 G 列
                Worksheet.Columns(col).HorizontalAlignment = -4108  # xlCenter
                Worksheet.Columns(col).VerticalAlignment = -4108  # xlCenter
        Worksheet.Columns("G").VerticalAlignment = -4108  # xlCenter

        # 设置自定义边距
        margin_points = 0.9 * 28.35  # 厘米转为点
        Worksheet.PageSetup.TopMargin = margin_points
        Worksheet.PageSetup.BottomMargin = margin_points

        # 删除 H 列中非空的行
        for i in range(Worksheet.UsedRange.Rows.Count, 1, -1):
            if Worksheet.Cells(i, 8).Value is not None:
                Worksheet.Rows(i).Delete()

        # 添加边框
        thin_border = 2  # xlThin
        for row in range(1, Worksheet.UsedRange.Rows.Count + 1):
            if Worksheet.Cells(row, 1).Value is not None:
                for col in range(1, 10):
                    Worksheet.Cells(row, col).Borders.Weight = thin_border

        # 添加隔行背景颜色
        for row in range(1, Worksheet.UsedRange.Rows.Count + 1):
            if row % 2 == 0:
                Worksheet.Rows(row).Interior.Color = 0xFFFFFF  # 白色
            else:
                Worksheet.Rows(row).Interior.Color = 0xF2F2F2  # 灰色

        # 添加页码到页脚
        Worksheet.PageSetup.CenterFooter = "第&P页, 共&N页"

        # 设置打印区域
        Worksheet.PageSetup.PrintArea = f'A1:I{Worksheet.UsedRange.Rows.Count}'

        # 设置为横向打印，并调整为 A4 尺寸
        Worksheet.PageSetup.Orientation = 2  # xlLandscape
        Worksheet.PageSetup.PaperSize = 9  # xlPaperA4
        Worksheet.PageSetup.Zoom = False  # 关闭缩放
        Worksheet.PageSetup.FitToPagesWide = 1
        Worksheet.PageSetup.FitToPagesTall = False

        # 保存更改
        Workbook.Save()
        print("Excel 文件已保存。")

        # 设置 PDF 文件名和路径
        date_string = datetime.now().strftime('%Y-%m-%d-%H%M')
        output_filename = f"{date_string}-许可.pdf"
        output_path = os.path.join(jbc_folder, output_filename)

        # 等待文件系统稳定
        time.sleep(2)

        # 导出为 PDF
        try:
            Workbook.ExportAsFixedFormat(0, output_path)
            print(f"PDF 文件已保存: {output_path}")
        except Exception as e:
            print(f"Error exporting to PDF: {e}")

        # 关闭 Workbook
        Workbook.Close(SaveChanges=False)

    except Exception as e:
        # 记录错误日志并打印错误信息
        logging.exception("An error occurred:")
        print(f"An error occurred: {str(e).encode('utf-8', errors='ignore').decode('utf-8')}")

    finally:
        # 确保 Excel 应用程序退出
        if Excel:
            Excel.Quit()
        print("Excel 进程已退出。")
                                                
def main():
    # Download CSV files
    csv_files = [
        download_csv("JUTB", "JUTp&HJKL2SJYjjuutt"),
        download_csv("UOFB", "EWQ&6qwe42B")
    ]

    # 过滤掉 None 值
    csv_files = [f for f in csv_files if f is not None]

    if not csv_files:
        print("CSV 文件下载失败。程序退出。")
        return

    # 处理 CSV 文件
    filtered_filename = process_csv_files(csv_files)

    # 获取单号列表
    single_numbers = get_single_numbers(filtered_filename)

    # 下载 UOF 文件
    uof_file = download_uof_file()

    if uof_file:
        try:
            # 读取下载的 UOF 文件数据并赋值给 uof_data
            print(f"读取 UOF 文件：{uof_file}")
            uof_data = pd.read_excel(uof_file, engine='openpyxl')
        except Exception as e:
            print(f"读取 UOF 文件时出错：{e}")
            return
    else:
        print("UOF 文件下载失败。程序退出。")
        return

    # 创建并格式化 Excel 文件
    excel_filename = create_excel_file(single_numbers, uof_data)
    format_excel_file(excel_filename)

if __name__ == "__main__":
    main()
