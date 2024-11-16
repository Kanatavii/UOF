import os
import re
import time
import logging
import glob
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import win32com.client as win32


def download_csv(username, password):
    download_path = os.path.expanduser('~\\public_downloads')  # 将文件保存到公共目录
    chrome_options = Options()
    prefs = {
        'download.default_directory': download_path,
        'download.prompt_for_download': False,
        'profile.default_content_settings.popups': 0,
        'directory_upgrade': True
    }
    chrome_options.add_experimental_option('prefs', prefs)
    # 可选：使用无头模式以加快执行速度
    # chrome_options.add_argument('--headless')
    # 添加无头模式选项
    chrome_options.add_argument('--headless')  # 运行在后台的无头模式
    chrome_options.add_argument('--disable-gpu')  # 禁用GPU加速
    chrome_options.add_argument('--no-sandbox')  # 仅在Linux上运行时需要
    chrome_options.add_argument('--window-size=1920x1080')  # 设置窗口大小
    
    driver = webdriver.Chrome(options=chrome_options)
    
    try:
        print(f"开始下载用户 {username} 的 CSV 文件")

        # 访问登录页面
        driver.get("http://www.ankcustoms.com/login.aspx")
        print("已打开登录页面")

        # 登录
        username_input = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "TextBox1"))
        )
        username_input.send_keys(username)
        print("已输入用户名")

        password_input = driver.find_element(By.ID, "TextBox2")
        password_input.send_keys(password)
        print("已输入密码")

        login_button = driver.find_element(By.ID, "Button1")
        login_button.click()
        print("已点击登录按钮")

        # 等待页面加载
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "rdoTrans"))
        )
        print("登录成功，页面已加载")

        # 选择“根据转运时间查询”单选按钮
        radio_button = driver.find_element(By.ID, "rdoTrans")
        radio_button.click()
        print("已选择根据转运时间查询")

        # 设置起始和结束日期时间
        start_datetime = (datetime.now() - timedelta(days=2)).strftime("%Y-%m-%d %H:%M:%S")
        end_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"开始时间：{start_datetime}，结束时间：{end_datetime}")

        # 定位输入框
        start_input = driver.find_element(By.ID, "_easyui_textbox_input5")
        end_input = driver.find_element(By.ID, "_easyui_textbox_input6")

        # 输入日期和时间
        start_input.clear()
        start_input.send_keys(start_datetime)
        print("已输入开始时间")

        end_input.clear()
        end_input.send_keys(end_datetime)
        print("已输入结束时间")

        # 等待页面处理输入
        time.sleep(2)

        # 点击“CSV下载”链接
        csv_download_link = driver.find_element(By.CSS_SELECTOR, "a.easyui-linkbutton[data-options*='icon-save']")
        csv_download_link.click()
        print("已点击 CSV 下载链接")

        # 等待下载完成
        print("等待下载完成...")
        wait_for_download_completion(download_path)
        print("下载完成")

        # 获取下载的 CSV 文件
        downloaded_file = get_latest_downloaded_file(download_path)
        print(f"下载的 CSV 文件：{downloaded_file}")

    except Exception as e:
        logging.exception(f"Error downloading CSV for user {username}: {e}")
        print(f"Error downloading CSV for user {username}: {e}")
        downloaded_file = None
    finally:
        print("关闭浏览器...")
        driver.quit()
        print("浏览器已关闭")

    return downloaded_file

def wait_for_download_completion(download_path, timeout=60):
    initial_csv_files = set(os.listdir(download_path))  # 获取下载之前的文件列表
    seconds = 0

    while seconds < timeout:
        time.sleep(1)
        current_files = set(os.listdir(download_path))  # 当前下载目录中的文件
        new_files = current_files - initial_csv_files   # 新下载的文件
        csv_files = [f for f in new_files if f.endswith('.csv')]  # 只取 .csv 文件

        if csv_files:
            print(f"下载完成的 CSV 文件：{csv_files}")
            return csv_files[0]  # 返回找到的第一个 CSV 文件
        else:
            print(f"等待 CSV 文件下载中...{seconds}秒")
            seconds += 1

    print("下载未在超时时间内完成")
    return None  # 超时后返回 None



def get_latest_downloaded_file(download_path):
    downloaded_files = glob.glob(os.path.join(download_path, "*.csv"))
    if not downloaded_files:
        raise ValueError("在下载文件夹中未找到 CSV 文件。")
    latest_file = max(downloaded_files, key=os.path.getctime)
    return latest_file


def process_csv_files(csv_files):
    dfs = [pd.read_csv(f, encoding='utf-8') for f in csv_files]
    merged_df = pd.concat(dfs, ignore_index=True).dropna()
    merged_df.iloc[:, 2] = merged_df.iloc[:, 2].astype(str).str.replace("-", "")
    filtered_df = merged_df[merged_df.iloc[:, 11].str.contains("NAKAMURA|GB", case=False, na=False)]

    filtered_filename = datetime.now().strftime("%Y-%m-%d_filtered.csv")
    filtered_df.to_csv(filtered_filename, index=False)
    print(f"Filtered CSV file saved as {filtered_filename}")
    return filtered_filename

def get_single_numbers(filtered_filename):
    filtered_data = pd.read_csv(filtered_filename)
    return filtered_data.iloc[:, 2].tolist()

def download_uof_file():
    download_path = os.path.expanduser('~\\public_downloads')  # 将文件保存到公共目录
    
     # 清除下载目录中旧的 UOF 文件
    clear_old_uof_files(download_path)
    
    chrome_options = Options()
    prefs = {
        'download.default_directory': download_path,
        'download.prompt_for_download': False,
        'profile.default_content_settings.popups': 0,
        'directory_upgrade': True,
        'safebrowsing.enabled': True
    }
    chrome_options.add_experimental_option('prefs', prefs)
    
    chrome_options.add_argument('--headless')  # 运行在后台的无头模式
    chrome_options.add_argument('--disable-gpu')  # 禁用GPU加速
    chrome_options.add_argument('--no-sandbox')  # 仅在Linux上运行时需要
    chrome_options.add_argument('--window-size=1920x1080')  # 设置窗口大小
    
    driver = webdriver.Chrome(options=chrome_options)

    try:
        driver.get("https://quickconnect.to/")

        # 输入 QuickConnect ID
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.ID, "input-id"))
        ).send_keys("uof-jp")
        driver.find_element(By.ID, "input-submit").click()

        # 登录
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.NAME, "username"))
        ).send_keys("anguri")
        driver.find_element(By.XPATH, "//div[contains(@class, 'login-btn-spinner-wrapper')]").click()

        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.NAME, "current-password"))
        ).send_keys("uofjpA-1")
        driver.find_element(By.XPATH, "//div[contains(@class, 'login-btn-spinner-wrapper')]").click()

        # 等待桌面加载完成
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.ID, "sds-desktop"))
        )
        print("登录成功，桌面已加载。")

        # 重试点击 File Station
        max_click_retries = 5
        click_retry_count = 0
        file_station_opened = False

        while not file_station_opened and click_retry_count < max_click_retries:
            try:
                file_station_icon = WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="sds-desktop-shortcut"]/div/li[1]/div[1]'))
                )
                driver.execute_script("arguments[0].click();", file_station_icon)
                file_station_opened = True
                print("已点击 File Station。")
            except Exception as e:
                click_retry_count += 1
                print(f"点击 File Station 失败。重试中...（{click_retry_count}/{max_click_retries}）")
                time.sleep(5)

        if not file_station_opened:
            print("多次尝试后未能打开 File Station。")
            return None

        # 切换到 File Station 窗口
        if len(driver.window_handles) > 1:
            driver.switch_to.window(driver.window_handles[-1])

        # 导航并下载文件
        action = webdriver.ActionChains(driver)
        uof_folder = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, "//div[contains(text(), 'UOF')]"))
        )
        action.double_click(uof_folder).perform()
        print("已进入 UOF 文件夹。")

        trans_data_folder = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, "//div[contains(text(),'转运数据')]"))
        )
        action.double_click(trans_data_folder).perform()
        print("已进入 转运数据 文件夹。")

        # 下载目标文件
        desired_filename = "UOF出入库汇总表.xlsx"
        file_element = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, f"//div[contains(text(), '{desired_filename}')]"))
        )
        action.double_click(file_element).perform()
        print("已开始下载 UOF 文件。")

        # 等待文件下载完成
        latest_uof_file = wait_for_uof_file_completion(download_path)
        if latest_uof_file:
            print(f"UOF 文件下载完成：{latest_uof_file}")
        else:
            print("UOF 文件下载失败。")
            return None

    except Exception as e:
        logging.exception(f"Error downloading UOF file: {e}")
        print(f"Error downloading UOF file: {e}")
        latest_uof_file = None
    finally:
        driver.quit()

    return latest_uof_file

def wait_for_uof_file_completion(download_path, timeout=120):
    """等待 UOF 文件 (.xlsx) 下载完成"""
    seconds = 0
    pattern = re.compile(r'UOF出入库汇总表(?: \((\d+)\))?\.xlsx')  # 匹配 UOF 出入库汇总表 文件名

    while seconds < timeout:
        time.sleep(1)
        current_files = os.listdir(download_path)

        # 查找符合条件的 .xlsx 文件，且不包括 .crdownload 文件
        uof_files = [f for f in current_files if f.endswith('.xlsx') and pattern.match(f)]

        if uof_files:
            # 根据创建时间找到最新的文件
            latest_uof_file = max(uof_files, key=lambda f: os.path.getctime(os.path.join(download_path, f)))
            print(f"找到最新的 UOF 文件：{latest_uof_file}")
            return os.path.join(download_path, latest_uof_file)

        print(f"等待 UOF 文件下载中... {seconds}秒")
        seconds += 1

    print("下载未在超时时间内完成")
    return None

def clear_old_uof_files(download_path):
    """清理下载目录中旧的 UOF 文件"""
    old_files = glob.glob(os.path.join(download_path, "UOF出入库汇总表*.xlsx"))
    for file in old_files:
        try:
            os.remove(file)
            print(f"已删除旧文件：{file}")
        except Exception as e:
            print(f"删除文件 {file} 时出错：{e}")

def create_excel_file(single_numbers, uof_data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # 设置列名和宽度
    column_names = ["许可时间", "回数", "送り状番号", "箱数", "转运公司", "转运备注", "现场用-函数对应", "入库时间", "取件地"]
    column_widths = [11, 4.63, 13.25, 4.63, 18, 25, 58.13, 19.88, 15]
    sheet.append(column_names)
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width

    current_date = datetime.now().strftime("%Y/%m/%d")

    # 填写数据
    for single_number in single_numbers:
        if pd.notnull(single_number):  # 确保 single_number 不是 NaN
            match = uof_data[uof_data['送り状番号'].astype(str).str.contains(str(single_number), na=False)]
        else:
            match = pd.DataFrame()  # 如果 single_number 是 NaN，则不匹配任何数据

        if not match.empty:
            matched_row = match.iloc[0]
            sheet.append([current_date, "", single_number, matched_row['箱数'], matched_row['转运公司'], matched_row['转运备注'], matched_row['现场用-函数对应'], matched_row['入库时间'], matched_row['取件地']])
        else:
            sheet.append([current_date, "", single_number, "", "", "", "", "", ""])

    # 保存 Excel 文件，修改命名规则
    date_string = datetime.now().strftime('%Y-%m-%d-%H%M')
    excel_filename = f"{date_string}.xlsx"
    workbook.save(excel_filename)
    print(f"Excel 文件已创建并保存: {excel_filename}")
    return excel_filename

def format_excel_file(excel_filename):
    try:
        
        # 创建Excel对象
        Excel = win32.DispatchEx("Excel.Application")
        Excel.Visible = False  # Excel在后台运行

        # 打开Excel文件
        filename = os.path.abspath(f"JBC{datetime.now().strftime('%Y%m%d%H%M')}.xlsx")
        Workbook = Excel.Workbooks.Open(filename)

        # 选择第一个sheet
        Worksheet = Workbook.Worksheets(1)


        # Set row height to 30
        for row in range(1, Worksheet.UsedRange.Rows.Count + 1):
            Worksheet.Rows(row).RowHeight = 30

        # Center and middle align columns A to F
        for col in range(1, 10):  # Columns A to H
            if col != 7:  # Skip column G (which is the 7th column)
                Worksheet.Columns(col).HorizontalAlignment = -4108  # xlCenter
                Worksheet.Columns(col).VerticalAlignment = -4108  # xlCenter

        # Middle align column G
        Worksheet.Columns("G").VerticalAlignment = -4108  # xlCenter

        # Set custom margins
        margin_top_bottom = 0.9  # In centimeters
        margin_points = margin_top_bottom * 28.35  # Convert centimeters to points
        Worksheet.PageSetup.TopMargin = margin_points
        Worksheet.PageSetup.BottomMargin = margin_points

        # 删除H列非空行
        for i in range(Worksheet.UsedRange.Rows.Count, 1, -1):
            if Worksheet.Cells(i, 8).Value is not None:
                Worksheet.Rows(i).Delete()

        # 添加全边框
        thin_border = 2
        for row in range(1, Worksheet.UsedRange.Rows.Count + 1):  # 从第1行开始
            if Worksheet.Cells(row, 1).Value is not None:
                for col in range(1, 10):  # 从A到H
                    Worksheet.Cells(row, col).Borders.Weight = thin_border

        # 设置打印区域
        Worksheet.PageSetup.PrintArea = 'A1:I{}'.format(Worksheet.UsedRange.Rows.Count)

        # 设置为横向
        Worksheet.PageSetup.Orientation = 2

        # 设置为A4纸
        Worksheet.PageSetup.PaperSize = 9

        # 设置为适应页面
        Worksheet.PageSetup.Zoom = False
        Worksheet.PageSetup.FitToPagesWide = 1
        Worksheet.PageSetup.FitToPagesTall = 1

        # 保存更改
        Workbook.Save()

        # 设置输出PDF文件名
        date_string = datetime.now().strftime('%Y%m%d%H%M')
        output_filename = f'JBC-{date_string}-许可.pdf'

        # 将Excel文件另存为PDF
        Workbook.ExportAsFixedFormat(0, os.path.join(os.path.dirname(filename), output_filename))

        # 退出Excel
        Workbook.Close(SaveChanges=False)
        Excel.Quit()

        print("done")

    except Exception as e:
        # Log any errors to the error.log file
        logging.exception("An error occurred:")
        print("An error occurred. Check the error.log file for details.")
                            
def main():
    # Download CSV files
    csv_files = [
        download_csv("JUTB", "JUTp&HJKL2SJYjjuutt"),
        download_csv("UOFB", "EWQ&6qwe42B")
    ]

    # 过滤掉 None 值
    csv_files = [f for f in csv_files if f is not None]

    if not csv_files:
        print("CSV 文件下载失败。程序退出。")
        return

    # 处理 CSV 文件
    filtered_filename = process_csv_files(csv_files)

    # 获取单号列表
    single_numbers = get_single_numbers(filtered_filename)

    # 下载 UOF 文件
    uof_file = download_uof_file()

    if uof_file:
        try:
            # 读取下载的 UOF 文件数据并赋值给 uof_data
            print(f"读取 UOF 文件：{uof_file}")
            uof_data = pd.read_excel(uof_file, engine='openpyxl')
        except Exception as e:
            print(f"读取 UOF 文件时出错：{e}")
            return
    else:
        print("UOF 文件下载失败。程序退出。")
        return

    # 创建并格式化 Excel 文件
    excel_filename = create_excel_file(single_numbers, uof_data)
    format_excel_file(excel_filename)

if __name__ == "__main__":
    main()
