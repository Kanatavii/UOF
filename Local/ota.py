import time
import logging
import re
import pandas as pd
from datetime import datetime, timedelta
import glob
import os
import openpyxl
import win32timezone
from selenium import webdriver
import win32com.client
from openpyxl.utils import get_column_letter
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains

# Configure the logging
logging.basicConfig(filename='error.log', level=logging.ERROR)

# 生成全局时间戳
timestamp = datetime.now().strftime('%Y%m%d%H%M')

# Put the code from your first script here
def download_csv(driver, username, password):
    # Navigate to the login page
    driver.get("http://www.ankcustoms.com/login.aspx")

    # Wait for the username input field to load
    username_input = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "TextBox1"))
    )
    username_input.send_keys(username)

    # Wait for the password input field to load
    password_input = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "TextBox2"))
    )
    password_input.send_keys(password)

    # Submit the login form
    login_button = driver.find_element(By.ID, "Button1")
    login_button.click()
                    
    # Find the element for "根据转运时间查询"
    radio_button = driver.find_element(By.ID, "rdoTrans")
    radio_button.click()

    # Get the position and size of the radio button
    radio_button_location = radio_button.location
    radio_button_size = radio_button.size

    # Calculate the target coordinates for the click
    target_x = radio_button_location["x"] - 150
    target_y = radio_button_location["y"] + (radio_button_size["height"] / 2)

    # Create an action chain
    actions = ActionChains(driver)

    # Move the mouse to the target coordinates and click
    actions.move_by_offset(target_x, target_y).click().perform()

    # Send Backspace key to delete 8 characters
    actions.send_keys(Keys.BACKSPACE * 8)

    # Get the current time
    current_time = time.strftime("%H:%M:%S")

    # Send the current time to the input field
    actions.send_keys(current_time).perform()


    # Find the element for "根据转运时间查询"
    radio_button = driver.find_element(By.ID, "rdoTrans")
    radio_button.click()

    # Get the position and size of the radio button
    radio_button_location = radio_button.location
    radio_button_size = radio_button.size

    # Calculate the target coordinates for the click
    offset_x = -300
    offset_y = radio_button.size["height"] / 2

    # Create an action chain
    actions = ActionChains(driver)

    # Move the mouse to the target coordinates and click
    actions.move_to_element(radio_button).move_by_offset(offset_x, offset_y).click().perform()

    # Send Backspace key to delete 8 characters
    actions.send_keys(Keys.BACKSPACE * 19)

    # Get the current time
    now = datetime.now()

    # Calculate the date for the day before yesterday
    day_before_yesterday = datetime.now() - timedelta(days=4)
    
    # Format the date as a string in the format "YYYY-M-D"
    date_str = day_before_yesterday.strftime("%Y-%m-%d")

    # Then you can send this time to the input field
    actions.send_keys(date_str).perform()

    # Wait for the selection to take effect (adjust sleep time as needed)
    time.sleep(1)
    
    # Find the element for "根据转运时间查询"
    radio_button = driver.find_element(By.ID, "rdoPermit")
    radio_button.click()

    # Find the element for CSV下载
    csv_download_link = driver.find_element(By.CSS_SELECTOR, "a.easyui-linkbutton[data-options*='icon-save']")

    # Click the CSV下载 link
    csv_download_link.click()
    
    time.sleep(3)
    
    # Get the downloaded CSV file
    downloaded_file = get_latest_downloaded_file()

    return downloaded_file


def get_latest_downloaded_file():
    downloads_folder = os.path.expanduser('~') + '/Downloads/'
    downloaded_files = glob.glob(downloads_folder + "*.csv")
    latest_file = max(downloaded_files, key=os.path.getctime)
    return latest_file

# Create WebDriver object
driver = webdriver.Chrome()

# Download the first CSV file
csv_file1 = download_csv(driver, "JUTB", "JUTp&HJKL2SJYjjuutt")

# Close the browser
driver.quit()

# Create WebDriver object
driver = webdriver.Chrome()

# Download the second CSV file
csv_file2 = download_csv(driver, "UOFB", "EWQ&6qwe42B")

# Close the browser
driver.quit()

# Merge the CSV files into a single DataFrame
df1 = pd.read_csv(csv_file1)
df2 = pd.read_csv(csv_file2)
merged_df = pd.concat([df1, df2], ignore_index=True)

# Create a new filename with the global timestamp
filename = f"{timestamp}.csv"

# Save the merged DataFrame to a new CSV file
merged_df.to_csv(filename, index=False)

# Load the merged CSV file
merged_df = pd.read_csv(filename)

# Replace "-" in the third column with empty string
merged_df.iloc[:, 2] = merged_df.iloc[:, 2].str.replace("-", "")

# Filter rows in the twelfth column that contain "NAKAMURA" or "GB" (case-insensitive)
filtered_df = merged_df[merged_df.iloc[:, 11].str.contains("NAKAMURA", case=False)]

# Create a new filename for the filtered DataFrame with the global timestamp
filtered_filename = f"{timestamp}_filtered.csv"

# Save the filtered DataFrame to a new CSV file
filtered_df.to_csv(filtered_filename, index=False)

print(f"Filtered CSV file saved as {filtered_filename}")

# Get the single numbers from filtered.csv
filtered_data = pd.read_csv(filtered_filename)
single_numbers = filtered_data.iloc[:, 2].tolist()

# 指定两个表格的文件路径
jjs_file = r"Z:\UOF\转运数据\JJS出入库汇总表.xlsx"
uof_file = r"Z:\UOF\转运数据\UOF出入库汇总表.xlsx"

# 读取两个表格
jjs_data = pd.read_excel(jjs_file)
uof_data = pd.read_excel(uof_file)

# 创建 Excel 工作簿
workbook = openpyxl.Workbook()
sheet = workbook.active

# 设置列名称
column_names = ["许可时间", "回数", "送り状番号", "箱数", "转运公司", "转运备注", "现场用-函数对应", "入库时间", "取件地"]
for i, column_name in enumerate(column_names, start=1):
    sheet.cell(row=1, column=i, value=column_name)

sheet.column_dimensions['A'].width = 11
sheet.column_dimensions['B'].width = 4.63
sheet.column_dimensions['C'].width = 13.25
sheet.column_dimensions['D'].width = 4.63
sheet.column_dimensions['E'].width = 18
sheet.column_dimensions['F'].width = 45
sheet.column_dimensions['G'].width = 58.13
sheet.column_dimensions['H'].width = 19.88
sheet.column_dimensions['I'].width = 15

# 设置日期格式
current_date = datetime.now().strftime("%Y/%m/%d")

# 从第二行开始填充数据
row = 2
for single_number in single_numbers:
    single_number_clean = str(single_number).strip().upper()

    # 填充基础数据
    sheet.cell(row=row, column=1, value=current_date)
    sheet.cell(row=row, column=2, value="")
    sheet.cell(row=row, column=3, value=single_number_clean)
    sheet.cell(row=row, column=4, value="")
    sheet.cell(row=row, column=5, value="")
    sheet.cell(row=row, column=6, value="")
    sheet.cell(row=row, column=7, value="")
    sheet.cell(row=row, column=8, value="")
    sheet.cell(row=row, column=9, value="")

    # 先在第一个表格 (JJS) 中查找
    order_data_jjs = jjs_data.loc[jjs_data["送り状番号"] == single_number_clean, ["箱数", "转运公司", "转运备注", "现场用-函数对应", "入库时间", "取件地"]]

    # 如果找到数据，则写入 Excel
    if not order_data_jjs.empty:
        vlookup_data = order_data_jjs.iloc[0].fillna("").tolist()

        # 写入数据到 Excel 文件
        for i, value in enumerate(vlookup_data, start=4):
            sheet.cell(row=row, column=i, value=value)

    # 然后在第二个表格 (UOF) 中查找
    order_data_uof = uof_data.loc[uof_data["送り状番号"] == single_number_clean, ["箱数", "转运公司", "转运备注", "现场用-函数对应", "入库时间", "取件地"]]

    # 如果找到数据，则写入 Excel
    if not order_data_uof.empty:
        vlookup_data = order_data_uof.iloc[0].fillna("").tolist()

        # 写入数据到 Excel 文件
        for i, value in enumerate(vlookup_data, start=4):
            sheet.cell(row=row, column=i, value=value)

    row += 1
    
# 筛选第九列，保留特定的选项
valid_options = ["山九", "山九-直取直送", "市川", "市川-直取直送", ""]

# 遍历每一行的第九列，删除不符合条件的行
for row_num in range(sheet.max_row, 1, -1):  # 从最后一行开始往上遍历
    cell_value = sheet.cell(row=row_num, column=9).value
    if cell_value not in valid_options:
        sheet.delete_rows(row_num)  # 删除整行

# 尝试保存并确保刷新文件数据
try:
    new_filename = rf"Z:\UOF\转运数据\许可\OTA{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
    workbook.save(new_filename)

except Exception as e:
    print(f"保存 Excel 文件时出现错误: {e}")
    
# 创建Excel对象
Excel = win32com.client.DispatchEx("Excel.Application")
Excel.Visible = False  # Excel在后台运行

# 打开Excel文件
Workbook = Excel.Workbooks.Open(new_filename)
df = pd.read_excel(new_filename)

# 选择第一个sheet
Worksheet = Workbook.Worksheets(1)

# Set row height to 30
for row in range(1, Worksheet.UsedRange.Rows.Count + 1):
    Worksheet.Rows(row).RowHeight = 30

# Center and middle align columns A to H except for G
for col in range(1, 10):  # Columns A to H
    if col != 7:  # Skip column G (which is the 7th column)
        Worksheet.Columns(col).HorizontalAlignment = -4108  # xlCenter
        Worksheet.Columns(col).VerticalAlignment = -4108  # xlCenter

# Middle align column G
Worksheet.Columns("G").VerticalAlignment = -4108  # xlCenter

# Set custom margins
margin_top_bottom = 0.9  # In centimeters
margin_points = margin_top_bottom * 28.35  # Convert centimeters to points
Worksheet.PageSetup.TopMargin = margin_points
Worksheet.PageSetup.BottomMargin = margin_points

# 删除H列非空行
for i in range(Worksheet.UsedRange.Rows.Count, 1, -1):
    if Worksheet.Cells(i, 8).Value is not None:
        Worksheet.Rows(i).Delete()

# 添加全边框
thin_border = 2
for row in range(1, Worksheet.UsedRange.Rows.Count + 1):  # 从第1行开始
    if Worksheet.Cells(row, 1).Value is not None:
        for col in range(1, 10):  # 从A到H
            Worksheet.Cells(row, col).Borders.Weight = thin_border
            
# 设置打印区域
Worksheet.PageSetup.PrintArea = 'A1:I{}'.format(Worksheet.UsedRange.Rows.Count)

# 设置为横向
Worksheet.PageSetup.Orientation = 2

# 设置为A4纸
Worksheet.PageSetup.PaperSize = 9

# 设置为适应页面
Worksheet.PageSetup.Zoom = False
Worksheet.PageSetup.FitToPagesWide = 1
Worksheet.PageSetup.FitToPagesTall = 1

# 保存更改
Workbook.Save()

# 设置输出PDF文件名，使用全局时间戳
output_filename = f'OTA-{timestamp}-许可.pdf'

# 将Excel文件另存为PDF
Workbook.ExportAsFixedFormat(0, os.path.join(os.path.dirname(new_filename), output_filename))

# 退出Excel
Workbook.Close(SaveChanges=False)
Excel.Quit()

print("入库单已完成")