import os
import time
import logging
import glob
import getpass
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import win32com.client

# Configure the logging
logging.basicConfig(filename='error.log', level=logging.ERROR)


def download_csv(username, password):
    download_path = os.path.expanduser('~\\Downloads')
    chrome_options = Options()
    prefs = {'download.default_directory': download_path}
    chrome_options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(options=chrome_options)

    try:
        # Navigate to the login page
        driver.get("http://www.ankcustoms.com/login.aspx")

        # Login
        username_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "TextBox1"))
        )
        username_input.send_keys(username)

        password_input = driver.find_element(By.ID, "TextBox2")
        password_input.send_keys(password)

        login_button = driver.find_element(By.ID, "Button1")
        login_button.click()

        # Wait for page to load
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "rdoTrans"))
        )

        # Select "根据转运时间查询" radio button
        radio_button = driver.find_element(By.ID, "rdoTrans")
        radio_button.click()

        # Set date and time
        # Replace 'startDate' and 'startTime' with the actual IDs of your date and time input fields
        day_before_yesterday = (datetime.now() - timedelta(days=2)).strftime("%Y-%m-%d")
        current_time = datetime.now().strftime("%H:%M:%S")

        date_input = driver.find_element(By.ID, "startDate")
        time_input = driver.find_element(By.ID, "startTime")

        date_input.clear()
        date_input.send_keys(day_before_yesterday)
        time_input.clear()
        time_input.send_keys(current_time)

        # Click the CSV下载 link
        csv_download_link = driver.find_element(By.CSS_SELECTOR, "a.easyui-linkbutton[data-options*='icon-save']")
        csv_download_link.click()

        # Wait for the download to complete
        wait_for_download_completion(download_path)

        # Get the downloaded CSV file
        downloaded_file = get_latest_downloaded_file(download_path)

    except Exception as e:
        logging.exception(f"Error downloading CSV for user {username}")
        downloaded_file = None
    finally:
        driver.quit()

    return downloaded_file


def wait_for_download_completion(download_path, timeout=60):
    seconds = 0
    while seconds < timeout:
        time.sleep(1)
        if any(filename.endswith('.crdownload') for filename in os.listdir(download_path)):
            seconds += 1
        else:
            return
    logging.error("Download did not complete within timeout period")


def get_latest_downloaded_file(download_path):
    downloaded_files = glob.glob(os.path.join(download_path, "*.csv"))
    if not downloaded_files:
        raise ValueError("No CSV files found in the Downloads folder.")
    latest_file = max(downloaded_files, key=os.path.getctime)
    return latest_file


def process_csv_files(csv_files):
    dfs = [pd.read_csv(f) for f in csv_files]
    merged_df = pd.concat(dfs, ignore_index=True).dropna()
    merged_df.iloc[:, 2] = merged_df.iloc[:, 2].str.replace("-", "")
    filtered_df = merged_df[merged_df.iloc[:, 11].str.contains("NAKAMURA|GB", case=False, na=False)]

    filtered_filename = datetime.now().strftime("%Y-%m-%d_filtered.csv")
    filtered_df.to_csv(filtered_filename, index=False)
    print(f"Filtered CSV file saved as {filtered_filename}")
    return filtered_filename


def get_single_numbers(filtered_filename):
    filtered_data = pd.read_csv(filtered_filename)
    return filtered_data.iloc[:, 2].tolist()


def download_uof_file():
    download_path = os.path.expanduser('~\\Downloads')
    chrome_options = Options()
    prefs = {'download.default_directory': download_path}
    chrome_options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(options=chrome_options)

    try:
        # Navigate to the URL
        driver.get("https://quickconnect.to/")

        # Enter QuickConnect ID
        input_box = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "input-id"))
        )
        input_box.send_keys("uof-jp")

        # Submit QuickConnect ID
        submit_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.ID, "input-submit"))
        )
        submit_button.click()

        # Login
        username_box = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.NAME, "username"))
        )
        username_box.send_keys("anguri")
        driver.find_element(By.XPATH, "//div[contains(@class, 'login-btn-spinner-wrapper')]").click()

        password_box = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.NAME, "current-password"))
        )
        password_box.send_keys("uofjpA-1")
        driver.find_element(By.XPATH, "//div[contains(@class, 'login-btn-spinner-wrapper')]").click()

        # Open File Station
        file_station_icon = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="sds-desktop-shortcut"]/div/li[1]/div[1]'))
        )
        driver.execute_script("arguments[0].click();", file_station_icon)

        # Navigate to the desired folder
        action = webdriver.ActionChains(driver)
        uof_folder = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//div[contains(text(), 'UOF')]"))
        )
        action.double_click(uof_folder).perform()

        trans_data_folder = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//div[contains(text(),'转运数据')]"))
        )
        action.double_click(trans_data_folder).perform()

        # Download the desired file
        desired_filename = "UOF出入库汇总表.xlsx"
        file_element = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, f"//div[contains(text(), '{desired_filename}')]"))
        )
        action.double_click(file_element).perform()

        # Wait for download to complete
        wait_for_download_completion(download_path)
        latest_uof_file = get_latest_downloaded_file(download_path)

    except Exception as e:
        logging.exception("Error downloading UOF file")
        latest_uof_file = None
    finally:
        driver.quit()

    return latest_uof_file


def create_excel_file(single_numbers, uof_data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set column names and widths
    column_names = ["许可时间", "回数", "送り状番号", "箱数", "转运公司", "转运备注", "现场用-函数对应", "入库时间", "取件地"]
    column_widths = [11, 4.63, 13.25, 4.63, 18, 25, 58.13, 19.88, 15]
    sheet.append(column_names)
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width

    current_date = datetime.now().strftime("%Y/%m/%d")

    # Fill data
    for single_number in single_numbers:
        sheet.append([current_date, "", single_number, "", "", "", "", "", ""])

    # Save initial Excel file
    excel_filename = "JBC{}.xlsx".format(datetime.now().strftime('%Y%m%d%H%M'))
    workbook.save(excel_filename)

    # Load Excel file into pandas for VLOOKUP-like operation
    df_excel = pd.read_excel(excel_filename)

    # Perform VLOOKUP
    for index, row in df_excel.iterrows():
        single_number = str(row['送り状番号'])
        match = uof_data[uof_data['送り状番号'].astype(str).str.contains(single_number, na=False)]
        if not match.empty:
            df_excel.loc[index, ['箱数', '转运公司', '转运备注', '现场用-函数对应', '入库时间', '取件地']] = match.iloc[0][['箱数', '转运公司', '转运备注', '现场用-函数对应', '入库时间', '取件地']].values

    # Save updated Excel file
    df_excel.to_excel(excel_filename, index=False)
    print("Excel 文件已创建并保存。")
    return excel_filename


def format_excel_file(excel_filename):
    try:
        Excel = win32com.client.DispatchEx("Excel.Application")
        Excel.Visible = False

        Workbook = Excel.Workbooks.Open(os.path.abspath(excel_filename))
        Worksheet = Workbook.Worksheets(1)

        # Set row height and alignment
        Worksheet.Rows.RowHeight = 30
        for col in range(1, 10):
            if col != 7:
                Worksheet.Columns(col).HorizontalAlignment = -4108  # xlCenter
                Worksheet.Columns(col).VerticalAlignment = -4108     # xlCenter
        Worksheet.Columns(7).VerticalAlignment = -4108  # Column G

        # Set margins
        margin_points = 0.9 * 28.35
        Worksheet.PageSetup.TopMargin = margin_points
        Worksheet.PageSetup.BottomMargin = margin_points

        # Delete rows where Column H is not empty
        last_row = Worksheet.UsedRange.Rows.Count
        for row in range(last_row, 1, -1):
            if Worksheet.Cells(row, 8).Value:
                Worksheet.Rows(row).Delete()

        # Add borders
        thin_border = 2  # xlThin
        for row in range(1, Worksheet.UsedRange.Rows.Count + 1):
            for col in range(1, 10):
                Worksheet.Cells(row, col).Borders.Weight = thin_border

        # Page setup
        Worksheet.PageSetup.PrintArea = Worksheet.UsedRange.Address
        Worksheet.PageSetup.Orientation = 2  # xlLandscape
        Worksheet.PageSetup.PaperSize = 9    # xlPaperA4
        Worksheet.PageSetup.Zoom = False
        Worksheet.PageSetup.FitToPagesWide = 1
        Worksheet.PageSetup.FitToPagesTall = 1

        # Save and export as PDF
        Workbook.Save()
        date_string = datetime.now().strftime('%Y%m%d%H%M')
        output_filename = f'JBC-{date_string}-许可.pdf'
        Workbook.ExportAsFixedFormat(0, os.path.join(os.path.dirname(excel_filename), output_filename))

        Workbook.Close(SaveChanges=False)
        Excel.Quit()
        print("Excel 文件已格式化并导出为 PDF。")

    except Exception as e:
        logging.exception("An error occurred while formatting Excel file:")
        print("An error occurred. Check the error.log file for details.")


def main():
    # Download CSV files
    csv_files = [
        download_csv("JUTB", "JUTp&HJKL2SJYjjuutt"),
        download_csv("UOFB", "EWQ&6qwe42B")
    ]

    # Process CSV files
    filtered_filename = process_csv_files(csv_files)

    # Get single numbers
    single_numbers = get_single_numbers(filtered_filename)

    # Download UOF file
    latest_uof_file = download_uof_file()
    if not latest_uof_file:
        print("Failed to download UOF file.")
        return

    uof_data = pd.read_excel(latest_uof_file)

    # Create and format Excel file
    excel_filename = create_excel_file(single_numbers, uof_data)
    format_excel_file(excel_filename)


if __name__ == "__main__":
    main()
