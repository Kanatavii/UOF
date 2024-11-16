import os
import re
import time
import logging
import glob
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import win32com.client as win32
from concurrent.futures import ThreadPoolExecutor

def download_csv(username, password):
    download_path = os.path.expanduser('~\\Downloads')  # 使用默认的下载路径
    chrome_options = Options()
    prefs = {
        'download.default_directory': download_path,  # 所有文件都存储在同一个下载目录
        'download.prompt_for_download': False,
        'profile.default_content_settings.popups': 0,
        'directory_upgrade': True
    }
    chrome_options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(options=chrome_options)

    try:
        print(f"开始下载用户 {username} 的 CSV 文件")

        driver.get("http://www.ankcustoms.com/login.aspx")
        print("已打开登录页面")

        username_input = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "TextBox1"))
        )
        username_input.send_keys(username)
        print("已输入用户名")

        password_input = driver.find_element(By.ID, "TextBox2")
        password_input.send_keys(password)
        print("已输入密码")

        login_button = driver.find_element(By.ID, "Button1")
        login_button.click()
        print("已点击登录按钮")

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "rdoTrans"))
        )
        print("登录成功，页面已加载")

        radio_button = driver.find_element(By.ID, "rdoTrans")
        radio_button.click()
        print("已选择根据转运时间查询")

        start_datetime = (datetime.now() - timedelta(days=2)).strftime("%Y-%m-%d %H:%M:%S")
        end_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"开始时间：{start_datetime}，结束时间：{end_datetime}")

        start_input = driver.find_element(By.ID, "_easyui_textbox_input5")
        end_input = driver.find_element(By.ID, "_easyui_textbox_input6")

        start_input.clear()
        start_input.send_keys(start_datetime)
        print("已输入开始时间")

        end_input.clear()
        end_input.send_keys(end_datetime)
        print("已输入结束时间")

        time.sleep(2)

        csv_download_link = driver.find_element(By.CSS_SELECTOR, "a.easyui-linkbutton[data-options*='icon-save']")
        csv_download_link.click()
        print(f"已点击 CSV 下载链接, 等待 {username} 文件下载完成...")

        # 等待下载完成
        downloaded_file = wait_for_download_completion(download_path)

        if downloaded_file:
            print(f"用户 {username} 的文件下载成功：{downloaded_file}")
        else:
            print(f"用户 {username} 的文件下载失败")

    except Exception as e:
        logging.exception(f"Error downloading CSV for user {username}: {e}")
        print(f"Error downloading CSV for user {username}: {e}")
    finally:
        print(f"关闭浏览器，用户 {username}")
        driver.quit()

    return downloaded_file

def wait_for_download_completion(download_path, timeout=180):
    initial_csv_files = set(os.listdir(download_path))  # 获取下载前的文件列表
    seconds = 0

    while seconds < timeout:
        time.sleep(1)
        current_files = set(os.listdir(download_path))  # 当前文件夹的文件列表
        new_files = current_files - initial_csv_files   # 下载的新文件
        csv_files = [f for f in new_files if f.endswith('.csv') or f.endswith('.crdownload')]  # 筛选出 CSV 或临时文件

        if csv_files:
            downloaded_file = os.path.join(download_path, csv_files[0])
            print(f"下载的文件：{downloaded_file}")

            # 如果文件是临时文件，继续等待
            if downloaded_file.endswith('.crdownload'):
                print(f"文件 {downloaded_file} 还在下载中，继续等待...")
                continue

            # 确保文件存在并已下载完成
            if not os.path.exists(downloaded_file):
                print(f"文件 {downloaded_file} 不存在，等待文件写入完成...")
                continue

            try:
                # 使用 Pandas 读取 CSV 文件的第一列第二行
                df = pd.read_csv(downloaded_file, header=None)
                user_type = df.iloc[1, 0].strip().lower()  # 获取第二行第一列的内容并转换为小写
                if user_type not in ['jutb', 'uofb']:
                    raise ValueError(f"未找到有效的用户类型: {user_type}")

                # 获取当前时间戳来保证唯一性
                timestamp = datetime.now().strftime("%Y%m%d%H%M%S")

                # 根据用户类型命名文件
                new_filename = f"{user_type}-{timestamp}.csv"
                new_file_path = os.path.join(download_path, new_filename)

                # 重命名文件
                os.rename(downloaded_file, new_file_path)
                print(f"下载完成的 CSV 文件已重命名为：{new_filename}")
                return new_file_path

            except Exception as e:
                print(f"读取文件时出错或解析用户类型失败: {e}")
                return None
        else:
            print(f"等待 CSV 文件下载中...{seconds}秒")
            seconds += 1

    print("下载未在超时时间内完成")
    return None

def download_csv_with_retry(username, password, max_retries=1, download_path="~\\Downloads"):
    download_path = os.path.expanduser(download_path)  # 展开路径
    retries = 0
    downloaded_file = None

    while retries <= max_retries:
        print(f"尝试第 {retries + 1} 次下载用户 {username} 的文件")
        downloaded_file = download_csv(username, password)

        if downloaded_file:
            # 下载完成后检查文件是否正确
            expected_file = find_latest_file(download_path, prefix=username.lower())
            if expected_file:
                print(f"用户 {username} 的文件下载成功：{expected_file}")
                return expected_file
            else:
                print(f"用户 {username} 的文件下载失败，第 {retries + 1} 次尝试失败")
        else:
            print(f"用户 {username} 的文件下载失败，第 {retries + 1} 次尝试失败")

        retries += 1  # 增加重试次数

    print(f"用户 {username} 的文件下载在 {max_retries} 次尝试后失败")
    return None

def find_latest_file(download_path, prefix):
    """查找指定前缀的最新下载文件"""
    try:
        # 获取以 prefix 开头的所有 CSV 文件
        files = [f for f in os.listdir(download_path) if f.startswith(prefix) and f.endswith('.csv')]
        if not files:
            return None

        # 根据文件创建时间排序，返回最新的文件
        latest_file = max(files, key=lambda f: os.path.getctime(os.path.join(download_path, f)))
        return os.path.join(download_path, latest_file)
    except Exception as e:
        print(f"查找 {prefix} 文件时出错: {e}")
        return None
    
def get_latest_downloaded_file(download_path):
    downloaded_files = glob.glob(os.path.join(download_path, "*.csv"))
    if not downloaded_files:
        raise ValueError("在下载文件夹中未找到 CSV 文件。")
    latest_file = max(downloaded_files, key=os.path.getctime)
    return latest_file

def process_csv_files(csv_files):
    dfs = [pd.read_csv(f, encoding='utf-8') for f in csv_files]
    merged_df = pd.concat(dfs, ignore_index=True).dropna()
    merged_df.iloc[:, 2] = merged_df.iloc[:, 2].astype(str).str.replace("-", "")
    filtered_df = merged_df[merged_df.iloc[:, 11].str.contains("NAKAMURA|GB", case=False, na=False)]

    filtered_filename = datetime.now().strftime("%Y-%m-%d_filtered.csv")
    filtered_df.to_csv(filtered_filename, index=False)
    print(f"Filtered CSV file saved as {filtered_filename}")
    return filtered_filename

def get_single_numbers(filtered_filename):
    filtered_data = pd.read_csv(filtered_filename)
    return filtered_data.iloc[:, 2].tolist()

def download_uof_file():
    download_path = os.path.expanduser('~\\Downloads')
    
    clear_old_uof_files(download_path)
    
    chrome_options = Options()
    prefs = {
        'download.default_directory': download_path,
        'download.prompt_for_download': False,
        'profile.default_content_settings.popups': 0,
        'directory_upgrade': True,
        'safebrowsing.enabled': True
    }
    chrome_options.add_experimental_option('prefs', prefs)

    
    driver = webdriver.Chrome(options=chrome_options)

    try:
        driver.get("https://quickconnect.to/")

        # 输入 QuickConnect ID
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.ID, "input-id"))
        ).send_keys("uof-jp")
        driver.find_element(By.ID, "input-submit").click()

        # 登录
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.NAME, "username"))
        ).send_keys("anguri")
        driver.find_element(By.XPATH, "//div[contains(@class, 'login-btn-spinner-wrapper')]").click()

        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.NAME, "current-password"))
        ).send_keys("uofjpA-1")
        driver.find_element(By.XPATH, "//div[contains(@class, 'login-btn-spinner-wrapper')]").click()

        # 等待桌面加载完成
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.ID, "sds-desktop"))
        )
        print("登录成功，桌面已加载。")

        # 重试点击 File Station
        max_click_retries = 5
        click_retry_count = 0
        file_station_opened = False

        while not file_station_opened and click_retry_count < max_click_retries:
            try:
                file_station_icon = WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="sds-desktop-shortcut"]/div/li[1]/div[1]'))
                )
                driver.execute_script("arguments[0].click();", file_station_icon)
                file_station_opened = True
                print("已点击 File Station。")
            except Exception as e:
                click_retry_count += 1
                print(f"点击 File Station 失败。重试中...（{click_retry_count}/{max_click_retries}）")
                time.sleep(5)

        if not file_station_opened:
            print("多次尝试后未能打开 File Station。")
            return None

        # 切换到 File Station 窗口
        if len(driver.window_handles) > 1:
            driver.switch_to.window(driver.window_handles[-1])

        # 导航并下载文件
        action = webdriver.ActionChains(driver)
        uof_folder = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, "//div[contains(text(), 'UOF')]"))
        )
        action.double_click(uof_folder).perform()
        print("已进入 UOF 文件夹。")

        trans_data_folder = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, "//div[contains(text(),'转运数据')]"))
        )
        action.double_click(trans_data_folder).perform()
        print("已进入 转运数据 文件夹。")

        # 下载目标文件
        desired_filename = "UOF出入库汇总表.xlsx"
        file_element = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, f"//div[contains(text(), '{desired_filename}')]"))
        )
        action.double_click(file_element).perform()
        print("已开始下载 UOF 文件。")

        # 等待文件下载完成
        latest_uof_file = wait_for_uof_file_completion(download_path)
        if latest_uof_file:
            print(f"UOF 文件下载完成：{latest_uof_file}")
        else:
            print("UOF 文件下载失败。")
            return None

    except Exception as e:
        logging.exception(f"Error downloading UOF file: {e}")
        print(f"Error downloading UOF file: {e}")
        latest_uof_file = None
    finally:
        driver.quit()

    return latest_uof_file

def wait_for_uof_file_completion(download_path, timeout=120):
    seconds = 0
    pattern = re.compile(r'UOF出入库汇总表(?: \((\d+)\))?\.xlsx')

    while seconds < timeout:
        time.sleep(1)
        current_files = os.listdir(download_path)
        uof_files = [f for f in current_files if f.endswith('.xlsx') and pattern.match(f)]

        if uof_files:
            latest_uof_file = max(uof_files, key=lambda f: os.path.getctime(os.path.join(download_path, f)))
            print(f"找到最新的 UOF 文件：{latest_uof_file}")
            return os.path.join(download_path, latest_uof_file)

        print(f"等待 UOF 文件下载中... {seconds}秒")
        seconds += 1

    print("下载未在超时时间内完成")
    return None

def clear_old_uof_files(download_path):
    old_files = glob.glob(os.path.join(download_path, "UOF出入库汇总表*.xlsx"))
    for file in old_files:
        try:
            os.remove(file)
            print(f"已删除旧文件：{file}")
        except Exception as e:
            print(f"删除文件 {file} 时出错：{e}")

def create_excel_file(single_numbers, uof_data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    column_names = ["许可时间", "回数", "送り状番号", "箱数", "转运公司", "转运备注", "现场用-函数对应", "入库时间", "取件地"]
    column_widths = [11, 4.63, 13.25, 4.63, 18, 25, 58.13, 19.88, 15]
    sheet.append(column_names)
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width

    current_date = datetime.now().strftime("%Y/%m/%d")

    for single_number in single_numbers:
        print(f"正在写入单号: {single_number}")
        match = uof_data[uof_data['送り状番号'].astype(str).str.contains(str(single_number), na=False)]
        if not match.empty:
            matched_row = match.iloc[0]
            sheet.append([current_date, "", single_number, matched_row['箱数'], matched_row['转运公司'], matched_row['转运备注'], matched_row['现场用-函数对应'], matched_row['入库时间'], matched_row['取件地']])
        else:
            sheet.append([current_date, "", single_number, "", "", "", "", "", ""])
        print(f"写入完成单号: {single_number}")

    date_string = datetime.now().strftime('%Y-%m-%d-%H%M')
    excel_filename = f"{date_string}.xlsx"
    file_path = os.path.join(os.getcwd(), excel_filename)
    workbook.save(file_path)
    print(f"Excel 文件已创建并保存: {file_path}")
    return file_path

def format_excel_file(excel_filename):
    try:
        
        # 创建Excel对象
        Excel = win32.DispatchEx("Excel.Application")
        Excel.Visible = False  # Excel在后台运行

        # 打开Excel文件
        Workbook = Excel.Workbooks.Open(os.path.abspath(excel_filename))  # 使用正确的文件路径

        # 选择第一个sheet
        Worksheet = Workbook.Worksheets(1)


        # Set row height to 30
        for row in range(1, Worksheet.UsedRange.Rows.Count + 1):
            Worksheet.Rows(row).RowHeight = 30

        # Center and middle align columns A to F
        for col in range(1, 10):  # Columns A to H
            if col != 7:  # Skip column G (which is the 7th column)
                Worksheet.Columns(col).HorizontalAlignment = -4108  # xlCenter
                Worksheet.Columns(col).VerticalAlignment = -4108  # xlCenter

        # Middle align column G
        Worksheet.Columns("G").VerticalAlignment = -4108  # xlCenter

        # Set custom margins
        margin_top_bottom = 0.9  # In centimeters
        margin_points = margin_top_bottom * 28.35  # Convert centimeters to points
        Worksheet.PageSetup.TopMargin = margin_points
        Worksheet.PageSetup.BottomMargin = margin_points

        # 删除H列非空行
        for i in range(Worksheet.UsedRange.Rows.Count, 1, -1):
            if Worksheet.Cells(i, 8).Value is not None:
                Worksheet.Rows(i).Delete()

        # 添加全边框
        thin_border = 2
        for row in range(1, Worksheet.UsedRange.Rows.Count + 1):  # 从第1行开始
            if Worksheet.Cells(row, 1).Value is not None:
                for col in range(1, 10):  # 从A到H
                    Worksheet.Cells(row, col).Borders.Weight = thin_border
                    
        # 添加隔行背景颜色
        for row in range(1, Worksheet.UsedRange.Rows.Count + 1):
            if row % 2 == 0:
                Worksheet.Rows(row).Interior.Color = 0xFFFFFF  # 白色
            else:
                Worksheet.Rows(row).Interior.Color = 0xF2F2F2  # 灰色

        # 添加页码
        Worksheet.PageSetup.CenterFooter = "&P页, 共&N页"

        # 设置打印区域
        Worksheet.PageSetup.PrintArea = 'A1:I{}'.format(Worksheet.UsedRange.Rows.Count)

        # 设置为横向
        Worksheet.PageSetup.Orientation = 2

        # 设置为A4纸
        Worksheet.PageSetup.PaperSize = 9

        # 设置为适应页面
        Worksheet.PageSetup.Zoom = False
        Worksheet.PageSetup.FitToPagesWide = 1
        Worksheet.PageSetup.FitToPagesTall = False

        # 保存更改
        Workbook.Save()

        # 设置 PDF 文件名
        date_string = datetime.now().strftime('%Y-%m-%d-%H%M')
        output_filename = f"{date_string}-许可.pdf"
        output_path = os.path.join(os.path.dirname(excel_filename), output_filename)

        # 将Excel文件另存为PDF
        Workbook.ExportAsFixedFormat(0, output_path)

        # 退出Excel
        Workbook.Close(SaveChanges=False)
        Excel.Quit()

        print("done")

    except Exception as e:
        # Log any errors to the error.log file
        logging.exception("An error occurred:")
        print("An error occurred. Check the error.log file for details.")

# 使用并行下载功能
def download_all_files():
    with ThreadPoolExecutor() as executor:
        futures = [
            executor.submit(download_csv, "JUTB", "JUTp&HJKL2SJYjjuutt"),
            executor.submit(download_csv, "UOFB", "EWQ&6qwe42B"),
            executor.submit(download_uof_file)
        ]
        csv_files = [f.result() for f in futures if f.result() is not None]

    return csv_files

def main():
    # 使用并行下载 CSV 文件和 UOF 文件
    csv_files = download_all_files()

    if not csv_files:
        print("CSV 文件下载失败。程序退出。")
        return

    # 处理 CSV 文件
    filtered_filename = process_csv_files(csv_files)

    # 获取单号列表
    single_numbers = get_single_numbers(filtered_filename)

    # 下载 UOF 文件
    uof_file = download_uof_file()

    if uof_file:
        uof_data = pd.read_excel(uof_file, engine='openpyxl')
    else:
        print("UOF 文件下载失败。程序退出。")
        return

    # 创建并格式化 Excel 文件
    excel_filename = create_excel_file(single_numbers, uof_data)
    format_excel_file(excel_filename)

if __name__ == "__main__":
    main()
